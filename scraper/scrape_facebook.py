#!/usr/bin/env python3
"""
GaukDarba — Facebook Groups job scraper

Scrapes Lithuanian job-posting Facebook groups using Playwright with a
persistent cookie session.  Exposes run() → {"jobs_found", "jobs_inserted", "error"}.

First run:  launches headful browser, waits for manual login, saves cookies.
Subsequent runs: loads cookies and runs headless.

Cookie file:  fb_cookies.json  (auto-created at runtime — not in VCS)
Profile dir:  fb_profile/       (auto-created at runtime — not in VCS)
Groups list:  facebook_groups.json

LOCAL TESTING MODE
------------------
Run directly without any env vars:
    python scrape_facebook.py

Results are written to  facebook_jobs_<timestamp>.xlsx  instead of Supabase.
Install extra dep if needed:  pip install openpyxl
"""

import json
import logging
import os
import random
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

SOURCE             = "facebook"
GROUPS_FILE        = Path(__file__).with_name("facebook_groups.json")
COOKIES_FILE       = Path(__file__).with_name("fb_cookies.json")
PROFILE_DIR        = Path(__file__).with_name("fb_profile")
MAX_SCROLLS        = 20
MAX_POST_AGE_DAYS  = 7
GROUP_COOLDOWN_HOURS = 6
BATCH_SIZE         = 50

# True when Supabase creds are present (Railway / production)
_SUPABASE_URL = os.environ.get("SUPABASE_URL")
_SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
LOCAL_MODE = not (_SUPABASE_URL and _SUPABASE_KEY)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)

LT_MONTHS = {
    "sausis": 1,  "sausio": 1,
    "vasaris": 2, "vasario": 2,
    "kovas": 3,   "kovo": 3,
    "balandis": 4,"balandžio": 4,
    "gegužė": 5,  "gegužės": 5,
    "birželis": 6,"birželio": 6,
    "liepa": 7,   "liepos": 7,
    "rugpjūtis": 8,"rugpjūčio": 8,
    "rugsėjis": 9,"rugsėjo": 9,
    "spalis": 10, "spalio": 10,
    "lapkritis": 11,"lapkričio": 11,
    "gruodis": 12,"gruodžio": 12,
}

JOB_KEYWORDS = [
    "ieškome", "darbo pasiūlymas", "samdome", "skelbimas",
    "darbas", "atlygis", "atlyginimas", " cv", "kandidatas",
    "etatas", "pozicija", "karjera", "darbuotojas", "darbuotojų",
    "įdarbinsime", "reikalingas", "reikalinga", "priimame",
]

GENERIC_FIRST_LINES = {
    "skubiai", "darbo pasiūlymas", "dėmesio", "informacija",
    "skelbimas", "darbas", "siūlome darbą", "siūlome",
}

LT_CITIES = [
    "Vilnius", "Kaunas", "Klaipėda", "Šiauliai", "Panevėžys",
    "Alytus", "Marijampolė", "Mažeikiai", "Jonava", "Utena",
    "Lietuva", "Vilniuje", "Kaune", "Klaipėdoje",
]

GROUP_CITY_HINTS = {
    "ivairūs_darbai_kaune": "Kaunas",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  [facebook]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("gaukdarba.facebook")

if LOCAL_MODE:
    log.info("*** LOCAL MODE — results will be saved to Excel, not Supabase ***")


# ── Emoji / noise stripping ───────────────────────────────────────────────────

_EMOJI_RE = re.compile(
    "[\U00010000-\U0010ffff"
    "\U0001F600-\U0001F64F"
    "\U0001F300-\U0001F5FF"
    "\U0001F680-\U0001F6FF"
    "\U0001F1E0-\U0001F1FF"
    "\u2600-\u26FF\u2700-\u27BF]+",
    flags=re.UNICODE,
)
_NOISE_RE = re.compile(r"[!*#•▶►▪▸]+")


def strip_noise(text: str) -> str:
    text = _EMOJI_RE.sub("", text)
    text = _NOISE_RE.sub("", text)
    return " ".join(text.split()).strip()


# ── Timestamp parsing ─────────────────────────────────────────────────────────

def parse_fb_timestamp(raw: str, now: datetime) -> datetime | None:
    raw = raw.strip().lower()

    m = re.match(r"(\d+)\s*min", raw)
    if m:
        return now - timedelta(minutes=int(m.group(1)))

    m = re.match(r"(\d+)\s*val", raw)
    if m:
        return now - timedelta(hours=int(m.group(1)))

    m = re.match(r"^(\d+)\s*d\.$", raw)
    if m:
        return now - timedelta(days=int(m.group(1)))

    m = re.match(
        r"([a-ząčęėįšųūž]+)\s+(\d+)\s+d\.(?:,\s*(\d+):(\d+))?", raw
    )
    if m:
        month_name = m.group(1)
        day  = int(m.group(2))
        hour = int(m.group(3)) if m.group(3) else 0
        minute = int(m.group(4)) if m.group(4) else 0
        month = LT_MONTHS.get(month_name)
        if month:
            year = now.year
            try:
                dt = datetime(year, month, day, hour, minute, tzinfo=timezone.utc)
                if dt > now:
                    dt = dt.replace(year=year - 1)
                return dt
            except ValueError:
                pass

    if raw.startswith("vakar"):
        return now - timedelta(days=1)

    return None


def unix_to_utc(utime_str: str) -> datetime | None:
    try:
        return datetime.fromtimestamp(int(utime_str), tz=timezone.utc)
    except (ValueError, OSError):
        return None


# ── Field extraction ──────────────────────────────────────────────────────────

def extract_title(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return ""
    first = strip_noise(lines[0])
    if first.lower() in GENERIC_FIRST_LINES or len(first) < 15:
        second = strip_noise(lines[1]) if len(lines) > 1 else ""
        first = (first + " — " + second).strip(" —") if second else first
    return first[:200]


def extract_company(text: str, poster_name: str | None) -> str | None:
    m = re.search(
        r"(?:Įmonė|Darbdavys)\s*:\s*([^\n,]{3,60})", text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    m = re.search(
        r'\b((?:UAB|MB|AB|VšĮ)\s+"?[^"\n,]{2,50}"?)', text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip().strip('"')

    return poster_name or None


def extract_salary(text: str) -> str | None:
    patterns = [
        r"\d[\d\s]*[-–]\s*\d[\d\s]*\s*€(?:\s*/\s*(?:mėn|val|h))?",
        r"\d[\d\s]*\s*€\s*/\s*(?:mėn|val|h)\b",
        r"(?:nuo|iki)\s*\d[\d\s]*\s*€",
        r"\d[\d\s]{2,}\s*€",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return " ".join(m.group(0).split())
    return None


def extract_location(text: str, group_name: str) -> str | None:
    m = re.search(
        r"(?:Darbo vieta|Miestas|Vieta)\s*:\s*([^\n,]{3,60})", text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    for city in LT_CITIES:
        if re.search(r"\b" + re.escape(city) + r"\b", text, re.IGNORECASE):
            return city

    return GROUP_CITY_HINTS.get(group_name)


def is_job_post(text: str) -> bool:
    if len(text) < 80:
        return False
    tl = text.lower()
    return any(kw in tl for kw in JOB_KEYWORDS)


def extract_post_id(url: str) -> str | None:
    m = re.search(r"/posts/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/permalink/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"story_fbid=(\d+)", url)
    if m:
        return m.group(1)
    return None


# ── Supabase helpers (skipped in LOCAL_MODE) ──────────────────────────────────

def group_last_run(supabase, group_name: str) -> datetime | None:
    if supabase is None:
        return None
    source_key = f"facebook_{group_name}"
    try:
        res = (
            supabase.table("scraper_runs")
            .select("ended_at")
            .eq("source", source_key)
            .order("ended_at", desc=True)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if not rows:
            return None
        raw = rows[0]["ended_at"]
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception as exc:
        log.warning("Could not query scraper_runs for %s: %s", source_key, exc)
        return None


def existing_fb_ids(supabase) -> set[str]:
    if supabase is None:
        return set()
    try:
        rows = (
            supabase.table("raw_listings")
            .select("job_id")
            .eq("source", SOURCE)
            .execute()
            .data or []
        )
        return {r["job_id"] for r in rows}
    except Exception as exc:
        log.warning("Could not fetch existing job IDs: %s", exc)
        return set()


def upsert_jobs(supabase, jobs: list[dict]) -> int:
    if supabase is None:
        return len(jobs)   # local mode: pretend everything "inserted"
    inserted = 0
    for i in range(0, len(jobs), BATCH_SIZE):
        chunk = jobs[i : i + BATCH_SIZE]
        try:
            res = supabase.table("raw_listings").upsert(
                chunk,
                on_conflict="job_id",
                ignore_duplicates=True,
            ).execute()
            inserted += len(res.data or [])
        except Exception as exc:
            log.warning("Upsert error (chunk %d): %s", i, exc)
    return inserted


def log_group_run(supabase, group_name, started, ended, found, inserted, error):
    if supabase is None:
        return
    try:
        supabase.table("scraper_runs").insert({
            "source":        f"facebook_{group_name}",
            "started_at":    started.isoformat(),
            "ended_at":      ended.isoformat(),
            "jobs_found":    found,
            "jobs_inserted": inserted,
            "error":         error,
        }).execute()
    except Exception as exc:
        log.warning("Could not write scraper_run for facebook_%s: %s", group_name, exc)


# ── Excel output (local mode only) ───────────────────────────────────────────

def save_excel(all_jobs: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facebook Jobs"

    headers = ["job_id", "title", "company", "salary_raw", "location", "url", "group", "source"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1877F2")   # Facebook blue

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_n, job in enumerate(all_jobs, start=2):
        ws.cell(row=row_n, column=1, value=job.get("job_id"))
        ws.cell(row=row_n, column=2, value=job.get("title"))
        ws.cell(row=row_n, column=3, value=job.get("company"))
        ws.cell(row=row_n, column=4, value=job.get("salary_raw"))
        ws.cell(row=row_n, column=5, value=job.get("location"))
        url_cell = ws.cell(row=row_n, column=6, value=job.get("url"))
        url_cell.hyperlink = job.get("url") or ""
        url_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_n, column=7, value=job.get("group_name"))
        ws.cell(row=row_n, column=8, value=job.get("source"))

    # Auto-size columns
    col_widths = [22, 60, 35, 20, 20, 70, 25, 12]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(__file__).with_name(f"facebook_jobs_{ts}.xlsx")
    wb.save(out_path)
    return out_path


# ── Playwright browser management ─────────────────────────────────────────────

def _is_login_wall(page) -> bool:
    url = page.url
    if "/login" in url or "login.php" in url:
        return True
    try:
        if page.query_selector("#login_form") or \
           page.query_selector('[data-testid="royal_login_form"]'):
            return True
    except Exception:
        pass
    return False


def _do_login(playwright):
    log.info("No valid cookies — launching headful browser for manual login.")
    log.info("Please log in to Facebook in the browser window (120 s timeout)...")

    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context(user_agent=USER_AGENT)
    page = context.new_page()
    page.goto("https://www.facebook.com/", timeout=30_000)

    try:
        page.wait_for_selector('[aria-label="Facebook"]', timeout=120_000)
        log.info("Login detected — saving session.")
    except Exception:
        context.close()
        browser.close()
        raise RuntimeError("Facebook login timed out (120 s).")

    context.storage_state(path=str(COOKIES_FILE))
    log.info("Cookies saved to %s", COOKIES_FILE)
    return context


def _launch_context(playwright):
    PROFILE_DIR.mkdir(exist_ok=True)

    kwargs = dict(
        user_agent=USER_AGENT,
        viewport={"width": 1280, "height": 900},
        locale="lt-LT",
    )

    if COOKIES_FILE.exists():
        log.info("Loading existing cookies from %s", COOKIES_FILE)
        try:
            context = playwright.chromium.launch_persistent_context(
                str(PROFILE_DIR),
                headless=True,
                storage_state=str(COOKIES_FILE),
                **kwargs,
            )
            page = context.new_page()
            page.goto("https://www.facebook.com/", timeout=30_000, wait_until="domcontentloaded")
            if not _is_login_wall(page):
                page.close()
                log.info("Cookie session valid.")
                return context
            log.warning("Cookie session expired — re-authenticating.")
            page.close()
            context.close()
        except Exception as exc:
            log.warning("Failed to load persistent context: %s", exc)

    return _do_login(playwright)


# ── Per-group scraping ────────────────────────────────────────────────────────

def _get_post_text(post) -> str:
    for selector in [
        '[data-ad-comet-preview="message"]',
        '[data-ad-preview="message"]',
        '[dir="auto"]',
    ]:
        try:
            el = post.query_selector(selector)
            if el:
                txt = el.inner_text()
                if txt and len(txt.strip()) > 10:
                    return txt.strip()
        except Exception:
            pass
    return ""


def _get_post_timestamp(post, now: datetime) -> datetime | None:
    try:
        abbr = post.query_selector("abbr[data-utime]")
        if abbr:
            utime = abbr.get_attribute("data-utime")
            if utime:
                return unix_to_utc(utime)
    except Exception:
        pass

    for selector in ["a[role='link'] abbr", "abbr"]:
        try:
            abbr = post.query_selector(selector)
            if abbr:
                raw = abbr.get_attribute("title") or abbr.inner_text()
                if raw:
                    dt = parse_fb_timestamp(raw, now)
                    if dt:
                        return dt
        except Exception:
            pass
    return None


def _get_post_permalink(post, group_id: str) -> str | None:
    for selector in ['a[href*="/posts/"]', 'a[href*="story_fbid="]', 'a[href*="/permalink/"]']:
        try:
            link = post.query_selector(selector)
            if link:
                href = link.get_attribute("href")
                if href:
                    clean_url = href.split("?")[0].rstrip("/") + "/"
                    if not clean_url.startswith("http"):
                        clean_url = "https://www.facebook.com" + clean_url
                    return clean_url
        except Exception:
            pass
    return None


def _get_poster_name(post) -> str | None:
    for selector in ["h3 a", "h4 a", "strong a"]:
        try:
            el = post.query_selector(selector)
            if el:
                name = el.inner_text().strip()
                if name:
                    return name
        except Exception:
            pass
    return None


def scrape_group(page, group: dict, known_ids: set[str], now: datetime) -> list[dict]:
    group_id   = group["id"]
    group_name = group["name"]
    group_url  = f"https://www.facebook.com/groups/{group_id}/"
    cutoff     = now - timedelta(days=MAX_POST_AGE_DAYS)

    log.info("  Navigating to group %s (%s)", group_name, group_url)
    try:
        page.goto(group_url, timeout=30_000, wait_until="domcontentloaded")
    except Exception as exc:
        raise RuntimeError(f"Timeout loading group {group_name}: {exc}")

    if _is_login_wall(page):
        raise RuntimeError(f"Login wall hit on group {group_name} — skipping.")

    jobs: list[dict] = []
    seen_posts: set[str] = set()
    no_new_streak = 0

    for scroll_n in range(MAX_SCROLLS):
        time.sleep(random.uniform(1.5, 3.0))

        try:
            posts = page.query_selector_all('[role="article"]')
        except Exception:
            posts = []

        new_this_scroll = 0
        stop_early = False

        for post in posts:
            try:
                permalink = _get_post_permalink(post, group_id)
                post_id   = extract_post_id(permalink) if permalink else None

                if not post_id:
                    continue
                if post_id in seen_posts:
                    continue
                seen_posts.add(post_id)

                job_id = f"fb_{post_id}"

                post_dt = _get_post_timestamp(post, now)
                if post_dt and post_dt < cutoff:
                    log.info(
                        "    Post %s dated %s — older than %dd cutoff, stopping.",
                        post_id, post_dt.date(), MAX_POST_AGE_DAYS,
                    )
                    stop_early = True
                    break

                if job_id in known_ids:
                    continue

                new_this_scroll += 1

                text = _get_post_text(post)
                if not is_job_post(text):
                    continue

                poster     = _get_poster_name(post)
                title      = extract_title(text)
                company    = extract_company(text, poster)
                salary_raw = extract_salary(text)
                location   = extract_location(text, group_name)

                if not title:
                    continue

                post_url = (
                    permalink
                    or f"https://www.facebook.com/groups/{group_id}/posts/{post_id}/"
                )

                jobs.append({
                    "job_id":     job_id,
                    "title":      title,
                    "company":    company,
                    "salary_raw": salary_raw,
                    "location":   location,
                    "url":        post_url,
                    "source":     SOURCE,
                    "group_name": group_name,   # extra field, stripped before DB upsert
                })
                known_ids.add(job_id)

            except Exception as exc:
                log.warning("    Post parse error: %s", exc)

        if stop_early:
            break

        if new_this_scroll == 0:
            no_new_streak += 1
            if no_new_streak >= 3:
                log.info("    No new posts for %d consecutive scrolls — stopping.", no_new_streak)
                break
        else:
            no_new_streak = 0

        log.info("    Scroll %d/%d — %d new this scroll, %d total for group",
                 scroll_n + 1, MAX_SCROLLS, new_this_scroll, len(jobs))

        try:
            page.evaluate("window.scrollBy(0, window.innerHeight * 2)")
        except Exception:
            pass

    return jobs


def _strip_extra_fields(jobs: list[dict]) -> list[dict]:
    """Remove fields not in raw_listings schema before upserting."""
    keep = {"job_id", "title", "company", "salary_raw", "location", "url", "source"}
    return [{k: v for k, v in j.items() if k in keep} for j in jobs]


# ── Main run ──────────────────────────────────────────────────────────────────

def run() -> dict:
    log.info("Facebook scraper starting  (local_mode=%s)", LOCAL_MODE)

    supabase = None
    if not LOCAL_MODE:
        from supabase import create_client
        supabase = create_client(_SUPABASE_URL, _SUPABASE_KEY)

    try:
        groups = json.loads(GROUPS_FILE.read_text(encoding="utf-8"))
    except Exception as exc:
        return {"jobs_found": 0, "jobs_inserted": 0, "error": f"Could not load groups file: {exc}"}

    now             = datetime.now(timezone.utc)
    cooldown_cutoff = now - timedelta(hours=GROUP_COOLDOWN_HOURS)
    known_ids       = existing_fb_ids(supabase)
    log.info("Loaded %d existing Facebook job IDs from DB", len(known_ids))

    total_found    = 0
    total_inserted = 0
    first_error: str | None = None
    all_jobs: list[dict] = []   # collected for Excel in local mode

    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        try:
            context = _launch_context(pw)
        except Exception as exc:
            err = f"Could not launch browser: {exc}"
            log.error(err)
            return {"jobs_found": 0, "jobs_inserted": 0, "error": err}

        try:
            page = context.new_page()
            page.set_extra_http_headers({"Accept-Language": "lt-LT,lt;q=0.9,en;q=0.5"})

            for group in groups:
                group_name    = group["name"]
                group_started = datetime.now(timezone.utc)

                last = group_last_run(supabase, group_name)
                if last and last > cooldown_cutoff:
                    log.info("  Skipping %s — last run %s (< %dh ago)",
                             group_name, last.strftime("%H:%M UTC"), GROUP_COOLDOWN_HOURS)
                    continue

                log.info("── Scraping group: %s ──", group_name)
                group_error: str | None = None
                group_jobs: list[dict] = []

                try:
                    group_jobs = scrape_group(page, group, known_ids, now)
                except Exception as exc:
                    group_error = str(exc)
                    log.error("  Group %s failed: %s", group_name, exc)
                    if not first_error:
                        first_error = group_error

                group_found    = len(group_jobs)
                group_inserted = 0

                if group_jobs:
                    all_jobs.extend(group_jobs)
                    db_jobs        = _strip_extra_fields(group_jobs)
                    group_inserted = upsert_jobs(supabase, db_jobs)

                total_found    += group_found
                total_inserted += group_inserted

                group_ended = datetime.now(timezone.utc)
                log_group_run(
                    supabase, group_name,
                    group_started, group_ended,
                    group_found, group_inserted, group_error,
                )

                log.info("  facebook_%s ✓  found=%d inserted=%d%s",
                         group_name, group_found, group_inserted,
                         f"  error={group_error}" if group_error else "")

        finally:
            try:
                context.storage_state(path=str(COOKIES_FILE))
            except Exception:
                pass
            context.close()

    # ── Local mode: dump to Excel ─────────────────────────────────────────
    if LOCAL_MODE and all_jobs:
        try:
            out = save_excel(all_jobs)
            log.info("Excel saved → %s  (%d rows)", out.name, len(all_jobs))
        except ImportError:
            log.warning("openpyxl not installed — run: pip install openpyxl")
        except Exception as exc:
            log.warning("Could not save Excel: %s", exc)
    elif LOCAL_MODE:
        log.info("No job posts found — nothing to save.")

    log.info("Facebook scraper done: total found=%d inserted=%d", total_found, total_inserted)
    return {"jobs_found": total_found, "jobs_inserted": total_inserted, "error": first_error}


if __name__ == "__main__":
    result = run()
    print(result)
    sys.exit(0 if result["error"] is None else 1)
